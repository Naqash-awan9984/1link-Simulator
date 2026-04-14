from http.server import BaseHTTPRequestHandler, HTTPServer
try:
    from http.server import ThreadingHTTPServer
except ImportError:
    # Python < 3.7 compatibility
    from socketserver import ThreadingMixIn

    class ThreadingHTTPServer(ThreadingMixIn, HTTPServer):
        daemon_threads = True
import json
import os
import re
import secrets
import threading
import time
from datetime import datetime
from urllib.parse import parse_qs
#import time

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

LOG_DIR = "/opt/1link-SIMULATOR/logs"
os.makedirs(LOG_DIR, exist_ok=True)
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
# If log directory creation fails (e.g., on Windows), use local logs directory
if not os.path.exists(LOG_DIR):
    LOG_DIR = os.path.join(BASE_DIR, "logs")
    os.makedirs(LOG_DIR, exist_ok=True)
FETCH_EXCEL_PATH = os.path.join(BASE_DIR, "Book16.xlsx")
FETCH_JSON_PATH  = os.path.join(BASE_DIR, "fetch_mappings.json")
DELAY_CONFIG_PATH = os.path.join(BASE_DIR, "delay_config.json")
TOKEN_STORE_PATH = os.path.join(BASE_DIR, "token_store.json")
OAUTH_TOKEN_PATH = "/oauth2/token"
OAUTH_GRANT_TYPE = "client_credentials"
OAUTH_VALID_CLIENTS = {
    ("c5a6f9d1895646dc83f4209eb2c8d770", "yO1lN5iF5jS4qY2lD8wW5hY3mS2wA0gT7fC7hM4qA7nX3sB5wK"),
    ("ac640fe07e809eeabe9b6120caa63c4e", "c2e7c81cd74a44c3c65cf6c0905ca4f0"),
}
OAUTH_SCOPE = "1LinkApi"
TOKEN_TTL_SECONDS = 24 * 60 * 60
TOKEN_STORE = {}

# Delay configuration (API-wise)
DEFAULT_DELAY_RULE = {
    "enabled": False,
    "delaySeconds": 40,
    "successResponsesBeforeDelay": 0,
    "delayCountBeforeSuccess": 0,
}
DEFAULT_DELAY_CONFIG = {
    "default": dict(DEFAULT_DELAY_RULE),
    "oauth": dict(DEFAULT_DELAY_RULE),
    "fetch": dict(DEFAULT_DELAY_RULE),
    "ibft": dict(DEFAULT_DELAY_RULE),
    "mpayTrace": dict(DEFAULT_DELAY_RULE),
}
DELAY_CONFIG = {key: dict(value) for key, value in DEFAULT_DELAY_CONFIG.items()}
DELAY_STATE = {
    "default": {"success_count": 0, "delay_count": 0},
    "oauth": {"success_count": 0, "delay_count": 0},
    "fetch": {"success_count": 0, "delay_count": 0},
    "ibft": {"success_count": 0, "delay_count": 0},
    "mpayTrace": {"success_count": 0, "delay_count": 0},
}
DELAY_CONFIG_MTIME = None

# IBFT RRN tracker (unique per day)
# Structure: {rrn: {"delay_count": int, "succeeded": bool, "response_pending": bool}}
RRN_TRACKER = {}
RRN_TRACKER_DATE = datetime.now().date()
RRN_TRACKER_LOCK = threading.Lock()


def log(msg):
    logfile = f"{LOG_DIR}/simulator_{datetime.now().strftime('%Y-%m-%d')}.log"
    with open(logfile,"a") as f:
        f.write(f"{datetime.now()} | {msg}\n")


def _load_delay_config():
    """Load delay configuration from file or use defaults."""
    def _rule_from_source(source):
        base = dict(DEFAULT_DELAY_RULE)
        if isinstance(source, dict):
            if "enabled" in source:
                base["enabled"] = bool(source.get("enabled"))
            if "delaySeconds" in source:
                try:
                    base["delaySeconds"] = float(source.get("delaySeconds"))
                except Exception:
                    pass
            if "successResponsesBeforeDelay" in source:
                try:
                    base["successResponsesBeforeDelay"] = int(source.get("successResponsesBeforeDelay"))
                except Exception:
                    pass
            if "delayCountBeforeSuccess" in source:
                try:
                    base["delayCountBeforeSuccess"] = int(source.get("delayCountBeforeSuccess"))
                except Exception:
                    pass 
        return base

    def _normalized_config(raw):
        # Backward compatibility: old flat shape applies to all APIs.
        if isinstance(raw, dict) and "enabled" in raw:
            flat_rule = _rule_from_source(raw)
            return {
                "default": dict(flat_rule),
                "oauth": dict(flat_rule),
                "fetch": dict(flat_rule),
                "ibft": dict(flat_rule),
                "mpayTrace": dict(flat_rule),
            }

        normalized = {key: dict(value) for key, value in DEFAULT_DELAY_CONFIG.items()}
        if isinstance(raw, dict):
            normalized["default"] = _rule_from_source(raw.get("default"))
            normalized["oauth"] = _rule_from_source(raw.get("oauth"))
            normalized["fetch"] = _rule_from_source(raw.get("fetch"))
            normalized["ibft"] = _rule_from_source(raw.get("ibft"))
            normalized["mpayTrace"] = _rule_from_source(raw.get("mpayTrace"))
        return normalized

    if os.path.exists(DELAY_CONFIG_PATH):
        try:
            with open(DELAY_CONFIG_PATH, "r") as f:
                config = json.load(f)
            config = _normalized_config(config)
            log(f"Delay config loaded: {config}")
            return config
        except Exception as exc:
            log(f"Failed to load delay_config.json: {exc}. Using defaults.")
    return {key: dict(value) for key, value in DEFAULT_DELAY_CONFIG.items()}


def _get_delay_bucket(request_path):
    if _is_oauth_token_path(request_path):
        return "oauth"
    if request_path == "/onelink/production/path-1":
        return "fetch"
    if request_path.endswith("/funds-transfer-rest-service/path-1"):
        return "ibft"
    if request_path == "/one-link/mpay/trace/IBFT":
        return "mpayTrace"
    return "default"


def _apply_api_delay(request_path):
    bucket = _get_delay_bucket(request_path)
    # IBFT delays are managed per-RRN in do_POST; skip here to avoid double-delay.
    if bucket == "ibft":
        return
    rule = DELAY_CONFIG.get(bucket, DELAY_CONFIG.get("default", DEFAULT_DELAY_RULE))
    state = DELAY_STATE.setdefault(bucket, {"success_count": 0, "delay_count": 0})

    if not rule.get("enabled", False):
        return

    delay_sec = float(rule.get("delaySeconds", 40))
    success_before_delay = int(rule.get("successResponsesBeforeDelay", 0))
    delay_count_needed = int(rule.get("delayCountBeforeSuccess", 0))

    should_delay = False
    if state["success_count"] >= success_before_delay:
        if delay_count_needed == 0:
            should_delay = True
        elif state["delay_count"] < delay_count_needed:
            should_delay = True

    if should_delay:
        log(f"Applying {bucket} delay: {delay_sec}s")
        time.sleep(delay_sec)
        state["delay_count"] += 1


def _mark_api_success(request_path):
    bucket = _get_delay_bucket(request_path)
    rule = DELAY_CONFIG.get(bucket, DELAY_CONFIG.get("default", DEFAULT_DELAY_RULE))
    if rule.get("enabled", False):
        state = DELAY_STATE.setdefault(bucket, {"success_count": 0, "delay_count": 0})
        state["success_count"] += 1


def _refresh_delay_config_if_changed():
    global DELAY_CONFIG, DELAY_CONFIG_MTIME, DELAY_STATE
    try:
        mtime = os.path.getmtime(DELAY_CONFIG_PATH)
    except OSError:
        return

    if DELAY_CONFIG_MTIME is None:
        DELAY_CONFIG_MTIME = mtime
        return

    if mtime != DELAY_CONFIG_MTIME:
        DELAY_CONFIG = _load_delay_config()
        DELAY_STATE = {
            "default": {"success_count": 0, "delay_count": 0},
            "oauth": {"success_count": 0, "delay_count": 0},
            "fetch": {"success_count": 0, "delay_count": 0},
            "ibft": {"success_count": 0, "delay_count": 0},
            "mpayTrace": {"success_count": 0, "delay_count": 0},
        }
        DELAY_CONFIG_MTIME = mtime
        log("Delay config reloaded due to file change; delay counters reset")


def _reset_rrn_tracker_if_needed():
    global RRN_TRACKER_DATE, RRN_TRACKER
    today = datetime.now().date()
    if today != RRN_TRACKER_DATE:
        with RRN_TRACKER_LOCK:
            if today != RRN_TRACKER_DATE:
                RRN_TRACKER = {}
                RRN_TRACKER_DATE = today
                log(f"RRN tracker reset for new day: {today}")


def _normalize_rrn(rrn_value):
    return str(rrn_value or "").strip()


def _get_rrn_decision(rrn_value, delay_count_needed):
    """
    Atomically determines what to do with an IBFT RRN.
    Returns (decision, attempt_num) where decision is:
      "delay"     - first request owns the transaction; delay its success response
      "succeed"   - process and return success immediately
      "duplicate" - transaction already owned/processed; return DUPLICATE TRANSACTION
    delay_count_needed: when > 0, IBFT success response is delayed for the first request.
    """
    rrn = _normalize_rrn(rrn_value)
    if not rrn:
        return ("succeed", 0)
    _reset_rrn_tracker_if_needed()
    with RRN_TRACKER_LOCK:
        entry = RRN_TRACKER.get(rrn)
        if entry is None:
            if delay_count_needed <= 0:
                RRN_TRACKER[rrn] = {"delay_count": 0, "succeeded": True, "response_pending": False}
                return ("succeed", 0)
            RRN_TRACKER[rrn] = {"delay_count": 1, "succeeded": True, "response_pending": True}
            return ("delay", 1)
        if entry.get("response_pending") or entry.get("succeeded"):
            return ("duplicate", 0)
        entry["succeeded"] = True
        return ("duplicate", 0)


def _complete_delayed_rrn_response(rrn_value):
    """Mark the first delayed IBFT response as no longer pending."""
    rrn = _normalize_rrn(rrn_value)
    if not rrn:
        return
    _reset_rrn_tracker_if_needed()
    with RRN_TRACKER_LOCK:
        entry = RRN_TRACKER.get(rrn)
        if entry is None:
            RRN_TRACKER[rrn] = {"delay_count": 1, "succeeded": True, "response_pending": False}
            return
        entry["succeeded"] = True
        entry["response_pending"] = False


def _normalize_header(value):
    return re.sub(r"[^a-z0-9]", "", str(value or "").strip().lower())


def _clean_imd(imd_value):
    return re.sub(r"\D", "", str(imd_value or ""))


def _clean_account(account_value):
    return re.sub(r"\s+", "", str(account_value or "")).upper()


def _find_column(header_map, candidates):
    for name in candidates:
        col_index = header_map.get(name)
        if col_index is not None:
            return col_index
    return None


def _parse_json_cell(value):
    if isinstance(value, dict):
        return value
    if value is None:
        return None

    text = str(value).strip()
    if not text:
        return None

    try:
        return json.loads(text)
    except Exception:
        return None


def _normalize_path(path_value):
    path = str(path_value or "").split("?", 1)[0].strip()
    # Some clients may send absolute-form URLs; keep only the path portion.
    if path.startswith("http://") or path.startswith("https://"):
        path = "/" + path.split("//", 1)[1].split("/", 1)[1] if "/" in path.split("//", 1)[1] else "/"
    return path.rstrip("/") or "/"


def _is_oauth_token_path(request_path):
    return request_path == OAUTH_TOKEN_PATH or request_path.endswith("/oauth2/token")


def _load_fetch_mappings():
    # Prefer JSON mapping (no external dependencies required).
    if os.path.exists(FETCH_JSON_PATH):
        try:
            with open(FETCH_JSON_PATH, "r") as f:
                mappings = json.load(f)
            log(f"Fetch mappings loaded from JSON: {len(mappings)} entries.")
            return mappings
        except Exception as exc:
            log(f"Failed to load fetch_mappings.json: {exc}. Trying Excel fallback.")

    # Fallback: load from Excel if openpyxl is available.
    if load_workbook is None:
        log("openpyxl is not installed and fetch_mappings.json not found. Using static fallback response.")
        return []

    if not os.path.exists(FETCH_EXCEL_PATH):
        log(f"Fetch mapping file not found: {FETCH_EXCEL_PATH}. Using static fallback response.")
        return []

    try:
        workbook = load_workbook(FETCH_EXCEL_PATH, data_only=True)
        sheet = workbook.active

        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []

        header_row = rows[0]
        header_map = {}
        for idx, header_value in enumerate(header_row):
            key = _normalize_header(header_value)
            if key and key not in header_map:
                header_map[key] = idx

        # Format 1: dedicated response columns (legacy/simple mapping)
        imd_col = _find_column(header_map, ["tobankimd", "imd", "toimd", "bankimd"])
        title_col = _find_column(header_map, ["accounttitleto", "accounttitle", "title", "customertitle"])
        branch_col = _find_column(header_map, ["branchnameto", "branchname", "branch"])
        bank_col = _find_column(header_map, ["bankname", "tobankname", "bank"])
        response_code_col = _find_column(header_map, ["responsecode", "respcode"])
        response_detail_col = _find_column(header_map, ["responsedetail", "responsedescription", "respdetail"])

        # Format 2: Book16 style with JSON columns
        tran_type_col = _find_column(header_map, ["trantype", "transactiontype"])
        client_req_col = _find_column(header_map, ["clientreq", "request", "clientrequest"])
        mpay_resp_col = _find_column(header_map, ["mpayresp", "response", "mparesp", "mpayresponse"])

        mappings = []

        # Parse Book16 rows where Fetch API request/response are stored as JSON.
        if tran_type_col is not None and client_req_col is not None and mpay_resp_col is not None:
            for row in rows[1:]:
                tran_type = str(row[tran_type_col]).strip().lower() if tran_type_col < len(row) and row[tran_type_col] is not None else ""
                if "title" not in tran_type or "fetch" not in tran_type:
                    continue

                request_obj = _parse_json_cell(row[client_req_col] if client_req_col < len(row) else None) or {}
                response_obj = _parse_json_cell(row[mpay_resp_col] if mpay_resp_col < len(row) else None) or {}

                title_fetch = {}
                if isinstance(response_obj, dict):
                    title_fetch = (response_obj.get("response") or {}).get("title-fetch") or {}

                imd = _clean_imd(
                    request_obj.get("ToBankIMD")
                    or title_fetch.get("ToBankIMD")
                    or ""
                )
                account_number_to = _clean_account(
                    request_obj.get("AccountNumberTo")
                    or title_fetch.get("AccountNumberTo")
                    or ""
                )

                if not imd:
                    continue

                mapping = {
                    "imd": imd,
                    "account_number_to": account_number_to,
                    "title": str(title_fetch.get("AccountTitleTo", "")).strip(),
                    "branch": str(title_fetch.get("BranchNameTo", "")).strip(),
                    "bank": str(title_fetch.get("BankName", "")).strip(),
                    "response_code": str(title_fetch.get("ResponseCode", "")).strip() or str((response_obj.get("response") or {}).get("response_code", "")).strip() or "00",
                    "response_detail": str(title_fetch.get("ResponseDetail", "")).strip() or str((response_obj.get("response") or {}).get("response_desc", "")).strip() or "PROCESSED OK",
                    "response_template": title_fetch if isinstance(title_fetch, dict) else {}
                }
                mappings.append(mapping)

        # Parse direct IMD-title mapping columns when available.
        if imd_col is not None:
            account_to_col = _find_column(header_map, ["accountnumberto", "accountto", "ibanmobileaccountnumber", "iban"])
            for row in rows[1:]:
                imd = _clean_imd(row[imd_col] if imd_col < len(row) else "")
                if not imd:
                    continue

                mappings.append({
                    "imd": imd,
                    "account_number_to": _clean_account(row[account_to_col]) if account_to_col is not None and account_to_col < len(row) and row[account_to_col] is not None else "",
                    "title": str(row[title_col]).strip() if title_col is not None and title_col < len(row) and row[title_col] is not None else "",
                    "branch": str(row[branch_col]).strip() if branch_col is not None and branch_col < len(row) and row[branch_col] is not None else "",
                    "bank": str(row[bank_col]).strip() if bank_col is not None and bank_col < len(row) and row[bank_col] is not None else "",
                    "response_code": str(row[response_code_col]).strip() if response_code_col is not None and response_code_col < len(row) and row[response_code_col] is not None else "00",
                    "response_detail": str(row[response_detail_col]).strip() if response_detail_col is not None and response_detail_col < len(row) and row[response_detail_col] is not None else "PROCESSED OK",
                    "response_template": {}
                })

        if not mappings:
            log("No usable Fetch mappings found in Excel file. Using static fallback response.")
            return []

        return mappings

    except Exception as exc:
        log(f"Failed to load fetch mappings from Excel: {exc}")
        return []


def _load_token_store():
    """Load persisted tokens from disk, discarding already-expired ones."""
    if not os.path.exists(TOKEN_STORE_PATH):
        return {}
    try:
        with open(TOKEN_STORE_PATH, "r") as f:
            raw = json.load(f)
        now_ts = int(time.time())
        return {tok: exp for tok, exp in raw.items() if isinstance(exp, (int, float)) and exp > now_ts}
    except Exception:
        return {}


def _save_token_store():
    """Persist current token store to disk."""
    try:
        with open(TOKEN_STORE_PATH, "w") as f:
            json.dump(TOKEN_STORE, f)
    except Exception as exc:
        log(f"Failed to save token store: {exc}")


FETCH_MAPPINGS = _load_fetch_mappings()
DELAY_CONFIG = _load_delay_config()
TOKEN_STORE.update(_load_token_store())
try:
    DELAY_CONFIG_MTIME = os.path.getmtime(DELAY_CONFIG_PATH)
except OSError:
    DELAY_CONFIG_MTIME = None


def _cleanup_expired_tokens(now_ts=None):
    now_value = now_ts if now_ts is not None else int(time.time())
    expired_tokens = [token for token, expiry in TOKEN_STORE.items() if expiry <= now_value]
    for token in expired_tokens:
        TOKEN_STORE.pop(token, None)
    if expired_tokens:
        _save_token_store()


def _issue_access_token():
    now_ts = int(time.time())
    _cleanup_expired_tokens(now_ts)
    token = secrets.token_urlsafe(48)
    TOKEN_STORE[token] = now_ts + TOKEN_TTL_SECONDS
    _save_token_store()
    return token, now_ts


def _parse_form_body(body):
    text = body or ""
    # Accept either standard form encoding (&) or comma-separated pairs used by some clients.
    if "&" not in text and "," in text:
        text = text.replace(",", "&")
    parsed = parse_qs(text, keep_blank_values=True)
    return {key: values[0] if values else "" for key, values in parsed.items()}


def _is_valid_oauth_client(client_id, client_secret):
    return (str(client_id).strip(), str(client_secret).strip()) in OAUTH_VALID_CLIENTS


def _get_bearer_token(authorization_header):
    header_value = str(authorization_header or "").strip()
    if not header_value.lower().startswith("bearer "):
        return ""
    return header_value[7:].strip()


def _is_authorized_request(headers):
    token = _get_bearer_token(headers.get("Authorization", ""))
    if not token:
        return False

    now_ts = int(time.time())
    _cleanup_expired_tokens(now_ts)
    expiry = TOKEN_STORE.get(token)
    if expiry is None:
        return False
    return expiry > now_ts


def _match_fetch_mapping(request_imd, request_account):
    req = _clean_imd(request_imd)
    req_account = _clean_account(request_account)
    if not req:
        return None

    # First preference: exact IMD + account match so repeated IMDs map to distinct cases.
    if req_account:
        for mapping in FETCH_MAPPINGS:
            value = mapping["imd"]
            mapped_account = _clean_account(mapping.get("account_number_to", ""))
            if not mapped_account:
                continue
            if (req == value or req.startswith(value) or value.startswith(req)) and req_account == mapped_account:
                return mapping

    for mapping in FETCH_MAPPINGS:
        value = mapping["imd"]
        if req == value or req.startswith(value) or value.startswith(req):
            return mapping

    return None


class OneLinkSimulator(BaseHTTPRequestHandler):

    def log_message(self, format, *args):
        # Suppress default stdout logging; we use our own log().
        pass

    def do_POST(self):
        response_data = None
        ibft_rrn_value = ""
        ibft_rrn_decision = ""
        try:
            _refresh_delay_config_if_changed()
            content_length = int(self.headers.get('Content-Length', 0))
            body = self.rfile.read(content_length).decode('utf-8', errors='replace')
            request_path = _normalize_path(self.path)

            try:
                request_json = json.loads(body)
            except Exception:
                request_json = {}

            # Per-RRN IBFT delay: delay the same RRN N times, then succeed on the (N+1)th attempt.
            if request_path.endswith("/funds-transfer-rest-service/path-1"):
                ibft_rule = DELAY_CONFIG.get("ibft", DEFAULT_DELAY_RULE)
                delay_count_needed = int(ibft_rule.get("delayCountBeforeSuccess", 0)) if ibft_rule.get("enabled", False) else 0
                rrn_value = request_json.get("RRN", "")
                ibft_rrn_value = rrn_value
                rrn_decision, attempt_num = _get_rrn_decision(rrn_value, delay_count_needed)
                ibft_rrn_decision = rrn_decision
                if rrn_decision == "duplicate":
                    log(f"IBFT duplicate detected for RRN: {_normalize_rrn(rrn_value)}")
                    log("================================")
                    log(f"Endpoint: {self.path}")
                    log(f"Request: {request_json}")
                    self.send_response(200)
                    self.send_header("Content-Type", "application/json")
                    self.end_headers()
                    response_data = json.dumps({
                        "ResponseCode": "10",
                        "PAN": str(request_json.get("PAN", "0000000000000") or "0000000000000"),
                        "Amount": str(request_json.get("Amount", "000001557800") or "000001557800"),
                        "TransmissionDateAndTime": str(request_json.get("TransmissionDateAndTime", "0403144546") or "0403144546"),
                        "STAN": str(request_json.get("STAN", "483505") or "483505"),
                        "Time": str(request_json.get("Time", "144546") or "144546"),
                        "Date": str(request_json.get("Date", "0403") or "0403"),
                        "DateSettlement": str(request_json.get("DateSettlement", "0404") or "0404"),
                        "RRN": str(request_json.get("RRN", "031445483505") or "031445483505"),
                        "AuthorizationIdentificationResponse": str(request_json.get("AuthorizationIdentificationResponse", "462911") or "462911"),
                        "ToBankIMD": str(request_json.get("ToBankIMD", "600648     ") or "600648     "),
                        "AccountNumberTo": str(request_json.get("AccountNumberTo", "PK57HABB0050797992069203") or "PK57HABB0050797992069203"),
                        "Reserved1": str(request_json.get("Reserved1", "") or ""),
                        "Reserved2": str(request_json.get("Reserved2", "") or ""),
                        "Reserved3": str(request_json.get("Reserved3", "") or ""),
                        "ResponseDetail": "DUPLICATE TRANSACTION"
                    })
                    self.wfile.write(response_data.encode())
                    log(f"Response: {response_data}")
                    log("================================\n")
                    return
                elif rrn_decision == "delay":
                    delay_sec = float(ibft_rule.get("delaySeconds", 40))
                    log(f"Applying ibft delay: {delay_sec}s (first request reserved success; retries will be duplicate)")
                    time.sleep(delay_sec)
                    _complete_delayed_rrn_response(rrn_value)
                    ibft_rrn_decision = "succeed"
                    # Client has timed out; fall through to send success response (broken pipe expected)

            log("================================")
            log(f"Endpoint: {self.path}")
            log(f"Request: {request_json}")

            # Apply endpoint-wise delay after request parsing/reservation.
            _apply_api_delay(request_path)

            if _is_oauth_token_path(request_path):
                form = _parse_form_body(body)
                grant_type = str(form.get("grant_type", "")).strip()
                client_id = str(
                    form.get("client_id", "")
                    or self.headers.get("X-IBM-Client-Id", "")
                ).strip()
                client_secret = str(
                    form.get("client_secret", "")
                    or self.headers.get("X-IBM-Client-Secret", "")
                ).strip()
                scope = str(form.get("scope", "")).strip()

                if (
                    grant_type != OAUTH_GRANT_TYPE
                    or not _is_valid_oauth_client(client_id, client_secret)
                    or scope != OAUTH_SCOPE
                ):
                    self.send_response(401)
                    self.send_header("Content-Type", "application/json")
                    self.end_headers()
                    response_data = json.dumps({
                        "error": "invalid_client",
                        "error_description": "Invalid OAuth credentials or grant parameters"
                    })
                    self.wfile.write(response_data.encode())
                    log(f"Response: {response_data}")
                    log("================================\n")
                    return

                access_token, consented_on = _issue_access_token()
                response_data = json.dumps({
                    "token_type": "Bearer",
                    "access_token": access_token,
                    "expires_in": TOKEN_TTL_SECONDS,
                    "consented_on": consented_on,
                    "scope": OAUTH_SCOPE
                })

                self.send_response(200)
                self.send_header("Content-Type", "application/json")
                self.end_headers()
                self.wfile.write(response_data.encode())
                log(f"Response: {response_data}")
                log("================================\n")
                return

            if not _is_authorized_request(self.headers):
                self.send_response(401)
                self.send_header("Content-Type", "application/json")
                self.end_headers()
                response_data = json.dumps({
                    "error": "invalid_token",
                    "error_description": "Provide a valid Bearer token in Authorization header"
                })
                self.wfile.write(response_data.encode())
                log(f"Response: {response_data}")
                log("================================\n")
                return

            # ------------------------------------------------
            # 1LINK FETCH API
            # ------------------------------------------------
            if request_path == "/onelink/production/path-1":

                mapping = _match_fetch_mapping(
                    request_json.get("ToBankIMD", ""),
                    request_json.get("AccountNumberTo", "")
                )

                account_title = "LUQMAN"
                branch_name = "HUB CHOWKI BRANCH"
                bank_name = "MEEZAN BANK LIMITED"
                response_code = "00"
                response_detail = "PROCESSED OK"
                to_bank_imd = "62787300000"

                if mapping:
                    account_title = mapping.get("title") or account_title
                    branch_name = mapping.get("branch") or branch_name
                    bank_name = mapping.get("bank") or bank_name
                    response_code = mapping.get("response_code") or response_code
                    response_detail = mapping.get("response_detail") or response_detail
                    to_bank_imd = mapping.get("imd") or to_bank_imd

                response = {
                    "ResponseCode": response_code,
                    "PAN": "0000000000000",
                    "Amount": "000005000000",
                    "TransmissionDateAndTime": request_json.get("TransmissionDateAndTime", ""),
                    "STAN": request_json.get("STAN", ""),
                    "Time": request_json.get("Time", ""),
                    "Date": request_json.get("Date", ""),
                    "DateSettlement": "0310",
                    "RRN": request_json.get("RRN", ""),
                    "AuthorizationIdentificationResponse": "672280",
                    "AccountNumberTo": request_json.get("AccountNumberTo", ""),
                    "ToBankIMD": to_bank_imd,
                    "AccountTitleTo": account_title,
                    "BranchNameTo": branch_name,
                    "BankName": bank_name,
                    "IBAN_MobileAccountNumber": request_json.get("AccountNumberTo", ""),
                    "ResponseDetail": response_detail
                }

                # If mapping provides a complete response template, use it exactly.
                if mapping and mapping.get("response_template"):
                    template = dict(mapping["response_template"])
                    response = {
                        "ResponseCode": str(template.get("ResponseCode", response_code) or response_code),
                        "PAN": str(template.get("PAN", "0000000000000") or "0000000000000"),
                        "Amount": str(template.get("Amount", "000005000000") or "000005000000"),
                        "TransmissionDateAndTime": str(template.get("TransmissionDateAndTime", request_json.get("TransmissionDateAndTime", ""))),
                        "STAN": str(template.get("STAN", request_json.get("STAN", ""))),
                        "Time": str(template.get("Time", request_json.get("Time", ""))),
                        "Date": str(template.get("Date", request_json.get("Date", ""))),
                        "DateSettlement": str(template.get("DateSettlement", "0310") or "0310"),
                        "RRN": str(template.get("RRN", request_json.get("RRN", ""))),
                        "AuthorizationIdentificationResponse": str(template.get("AuthorizationIdentificationResponse", "672280") or "672280"),
                        "AccountNumberTo": str(template.get("AccountNumberTo", request_json.get("AccountNumberTo", ""))),
                        "ToBankIMD": str(template.get("ToBankIMD", to_bank_imd) or to_bank_imd),
                        "AccountTitleTo": str(template.get("AccountTitleTo", account_title) or account_title),
                        "BranchNameTo": str(template.get("BranchNameTo", branch_name) or branch_name),
                        "BankName": str(template.get("BankName", bank_name) or bank_name),
                        "IBAN_MobileAccountNumber": str(template.get("IBAN_MobileAccountNumber", request_json.get("AccountNumberTo", ""))),
                        "ResponseDetail": str(template.get("ResponseDetail", response_detail) or response_detail)
                    }

                response_data = json.dumps(response)

            # ------------------------------------------------
            # 1LINK IBFT PAYMENT
            # ------------------------------------------------
            elif request_path.endswith("/funds-transfer-rest-service/path-1"):
                response = {
                    "ResponseCode": "00",
                    "PAN": str(request_json.get("PAN", "0000000000000") or "0000000000000"),
                    "Amount": str(request_json.get("Amount", "000000000100") or "000000000100"),
                    "TransmissionDateAndTime": str(request_json.get("TransmissionDateAndTime", "") or ""),
                    "STAN": str(request_json.get("STAN", "") or ""),
                    "Time": str(request_json.get("Time", "") or ""),
                    "Date": str(request_json.get("Date", "") or ""),
                    "DateSettlement": str(request_json.get("DateSettlement", "0411") or "0411"),
                    "RRN": str(request_json.get("RRN", "") or ""),
                    "AuthorizationIdentificationResponse": str(
                        request_json.get("AuthorizationIdentificationResponse", "092415") or "092415"
                    ),
                    "ToBankIMD": str(request_json.get("ToBankIMD", "998932     ") or "998932     "),
                    "AccountNumberTo": str(request_json.get("AccountNumberTo", "") or ""),
                    "Reserved1": str(request_json.get("Reserved1", "") or ""),
                    "Reserved2": str(request_json.get("Reserved2", "") or ""),
                    "Reserved3": str(request_json.get("Reserved3", "") or ""),
                    "ResponseDetail": "PROCESSED OK",
                }
                response_data = json.dumps(response)

            # ------------------------------------------------
            # MPAY TRACE API
            # ------------------------------------------------
            elif request_path == "/one-link/mpay/trace/IBFT":

                response = {
                    "response": {
                        "response_code": "0000",
                        "response_desc": "SUCCESS",
                        "retryCount": "1"
                    },
                    "info": {
                        "stan": request_json.get("ostan", ""),
                        "rrn": request_json.get("orrn", "")
                    }
                }

                response_data = json.dumps(response)

            else:
                self.send_response(404)
                self.send_header("Content-Type", "application/json")
                self.end_headers()
                self.wfile.write(json.dumps({"error": "Unknown endpoint"}).encode())
                log(f"Unknown Endpoint: {self.path}")
                log("================================\n")
                return

            self.send_response(200)
            self.send_header("Content-Type", "application/json")
            self.end_headers()
            self.wfile.write(response_data.encode())

            # Increment successful response count for this endpoint.
            _mark_api_success(request_path)

            log(f"Response: {response_data}")
            log("================================\n")

        except Exception as exc:
            # Broken pipe / connection reset means the client disconnected; not a server error.
            exc_str = str(exc)
            if any(phrase in exc_str for phrase in ("Broken pipe", "Connection reset", "[Errno 32]", "[Errno 104]", "[WinError 10053]", "[WinError 10054]")):
                if ibft_rrn_decision:
                    log(
                        f"Client disconnected before response could be sent: {exc} "
                        f"| ibft_rrn={_normalize_rrn(ibft_rrn_value)} | decision={ibft_rrn_decision}"
                    )
                else:
                    log(f"Client disconnected before response could be sent: {exc}")
                return
            log(f"ERROR in do_POST: {exc}")
            try:
                error_body = json.dumps({"error": str(exc)}).encode()
                self.send_response(500)
                self.send_header("Content-Type", "application/json")
                self.send_header("Content-Length", str(len(error_body)))
                self.end_headers()
                self.wfile.write(error_body)
            except Exception:
                pass


def run():
    port = 2011
    server = ThreadingHTTPServer(("0.0.0.0", port), OneLinkSimulator)

    log(f"BASE_DIR: {BASE_DIR}")
    log(f"fetch_mappings.json path: {FETCH_JSON_PATH}")
    log(f"fetch_mappings.json exists: {os.path.exists(FETCH_JSON_PATH)}")
    log(f"Fetch mappings loaded: {len(FETCH_MAPPINGS)}")
    log(f"Delay config: {DELAY_CONFIG}")
    print(f"1LINK Simulator running on port {port}")
    log("Simulator Started")

    server.serve_forever()


if __name__ == "__main__":
    run()
